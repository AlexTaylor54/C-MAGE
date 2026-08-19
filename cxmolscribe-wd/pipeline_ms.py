import sys
import torch
import rdkit.Chem as Chem
from rdkit.Chem import Draw
import os
import pandas as pd
import numpy as np
import time
import cv2
import argparse
from pathlib import Path
from openpyxl import load_workbook
from huggingface_hub import hf_hub_download


#Every path defaults to a location derived from this file rather than from the working directory
CXMS_DIR = Path(__file__).resolve().parent
DIS_DIR = CXMS_DIR / "DECIMER-Image-Segmentation"

def select_device(requested=None):
    """CUDA if present, otherwise CPU.

    MPS is deliberately not auto-selected. MolScribe predicts one image at a
    time, so the per-image transfer to the Apple GPU costs more than it saves:
    measured 34.6s on MPS against 7.0s on CPU for the same 12 structures, with
    identical predictions. Pass --device mps to use it anyway.
    """
    if requested:
        return torch.device(requested)
    if torch.cuda.is_available():
        return torch.device("cuda")
    return torch.device("cpu")


parser = argparse.ArgumentParser(
    description="Translate segmented structure images into CXSMILES, writing one "
                "spreadsheet without splitting on confidence. Alternative to folder_ms.py.")
parser.add_argument("--results-excel", type=Path, default=DIS_DIR / "DIS_CMAGE_results.xlsx",
                    help="Spreadsheet of segment paths written by stage 2 (DECIMER).")
parser.add_argument("--output-dir", type=Path, default=DIS_DIR,
                    help="Directory for the rendered structures and the completed spreadsheet.")
parser.add_argument("--device", default=os.environ.get("CMAGE_DEVICE"),
                    help="torch device: cuda, mps or cpu. Detected automatically "
                         "when not given. Also read from CMAGE_DEVICE.")
args = parser.parse_args()

if not args.results_excel.is_file():
    raise SystemExit(
        f"Stage 2 results not found: {args.results_excel}\n"
        "Run pipeline_dis.py first, or pass --results-excel.")

os.makedirs(args.output_dir, exist_ok=True)

#Establishes model path for being run in this script
model_path = hf_hub_download('yujieq/MolScribe', 'swin_base_char_aux_1m.pth')

#Makes a dataframe out of the excel file created from the DECIMER-Image-Segmentation proccessing
df = pd.read_excel(args.results_excel)

#Takes DECIMER-Image-Segmentation's output file paths and puts them in a list for CXMolScribe usage
file_paths = []
for paths in df["DIS Result File Paths"]:
    file_paths.append(paths)

#Loads the DECIMER-Image-Segmentations output excel as a workbook that can be edited
workbook = load_workbook(args.results_excel)
worksheet = workbook.active


from openpyxl.drawing.image import Image
from molscribe import MolScribe

#Establishes the model to be used for translation
device = select_device(args.device)
print("Device = " + str(device))
model = MolScribe(model_path, device)


#Creates the Appropriate Column Titles for CXMolScribe Output
worksheet.cell(row=1,column=2).value = "File Path"
worksheet.cell(row=1,column=4).value = "Image From File Path"
worksheet.cell(row=1, column=5).value = "Image From Predicted SMILES"
worksheet.cell(row=1, column=6).value = "Predicted SMILES"
worksheet.cell(row=1,column=7).value = "Nuances in Image"
worksheet.cell(row=1,column=8).value = "Holistic Molecule Interpretation"
worksheet.cell(row=1,column=9).value = "CXSMILES's Confidence Levels"
worksheet.cell(row=1,column=10).value = "Confidence Classification"
worksheet.cell(row=1,column=11).value = "Type of Mol"


#Creates a list for predicted CXSMILES strings to be placed into
smiles_prediction = []


#Confidence Classifcation Strings and Counters
correct_classification = "High Confidence of SMILES being correct"
discard_classification = "Discard this SMILES prediction"

starttime= time.time()

correct_counter = 0
discard_counter = 0

#Rendered structures go in their own folder rather than the working directory
image_folder = str(args.output_dir / "structure_images")
os.makedirs(image_folder, exist_ok=True)

row_value = 2
#For loop to run each file path through CXMolScribe and organizing the results
for digit,fps in enumerate(file_paths):
    
    #Makes graph/CXSMILES prediction of DECIMER-Image-Segmentation segmentation
    #Uses established CI-3.0 to sort molecules based on CXMolScribe confidence
    #Appropriate excel sheet variables are filled
    
    if digit == digit:
        try:
            prediction = model.predict_image_file(fps, return_atoms_bonds=True, return_confidence=True)
            smiles = prediction["smiles"]
            confidence = prediction["confidence"]

            #Retains the molecules figure connection within the naming of the CXSMILES translation
            cut_start = fps.index("_:") + 2
            cut_end = fps.index(".png")
            translation_identifier = fps[cut_start:cut_end]

            if confidence >= 0.8431 and str(smiles) != "<invalid>":
                worksheet.cell(row=row_value,column=10).value = correct_classification
                correct_counter += 1
            if confidence < 0.8431 and str(smiles) != "<invalid>":
                worksheet.cell(row=row_value,column=10).value = discard_classification
                discard_counter += 1
            if str(smiles) == "<invalid>":
                worksheet.cell(row=row_value,column=10).value = discard_classification
                discard_counter += 1
                
                #Appends molecule predictions to appropriate location on output spreadsheet
            worksheet.cell(row=row_value, column=9).value = confidence
            worksheet.cell(row=row_value, column=6).value = smiles
            smiles_prediction.append(smiles)

                #Creates an image of the ground truth DECIMER-Image-Segmentation output and organizes it into the excel sheet
            fp_image = Image(fps) 
            fp_image.width = 62
            fp_image.height = 48
            worksheet.add_image(fp_image, "D"+str(row_value))

                #Creates an image of the predicted SMILES string and organizes it into the excel sheet
            mol = Chem.MolFromSmiles(smiles)
            smiles_path = os.path.abspath(
                os.path.join(image_folder, str(translation_identifier) + ".png"))
            mol_image = Draw.MolsToImage([mol])
            mol_image.save(smiles_path)
            pred_smiles_image = Image(smiles_path)
            pred_smiles_image.width = 75
            pred_smiles_image.height = 70
            worksheet.add_image(pred_smiles_image, "E"+str(row_value))

                #Standardizes sizing of rows and columns in the excel sheet
            worksheet.row_dimensions[digit].height = 90
            worksheet.column_dimensions["D"].width = 70
            worksheet.column_dimensions["F"].width = 100
            worksheet.column_dimensions["G"].width = 40
            worksheet.column_dimensions["H"].width = 40
            worksheet.column_dimensions["I"].width = 50
            row_value +=1
        except:
            row_value += 1
            continue


endtime = time.time()
totaltime = endtime - starttime


print("Total CXMolScribe Time = " +str(totaltime)+ " seconds!")
print("High Confidence of Correct Counter = " +str(correct_counter))
print("Discard Counter = " +str(discard_counter))

#Saves Workbook as new excel file which now has all CXMolScribe results from above
workbook.save(args.output_dir / "Completed_CMAGE.xlsx")

print("Whole Pipeline Extracted!")
