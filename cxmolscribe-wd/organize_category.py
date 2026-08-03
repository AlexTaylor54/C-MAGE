import sys
import torch
import rdkit.Chem as Chem
from rdkit.Chem import Draw
import os
import pandas as pd
import numpy as np
import time
import cv2
from pathlib import Path
from openpyxl import load_workbook

"""
Once you have annotated your extracted dataset this file can be used to 
sort all of the different image categorizations into their own files for
further analysis and organization
"""

#Makes a dataframe out of annotated Excel file created from CXMolScribe results
#The file here should be your annotated Excel file
df = pd.read_excel("DECIMER-Image-Segmentation/BLANK.xlsx")#THIS WILL NEED TO CHANGE WITH WHAT YOU NAME ANNOTATED DATASET

#The following for loops will take the information present in the annotated Excel and place each columns values in a list
file_paths = []
for paths in df["DIS Result File Paths"]:
    file_paths.append(paths)

smiles_list = []
for smiles in df["Predicted SMILES"]:
    smiles_list.append(smiles)

nuance_list = []
for nuances in df["Nuances in Image"]:
    try:
        nuance_list.append(nuances)
        if nuances is None:
            nuances = ""
            nuance_list.append(nuances)

grade_list = []
for grades in df["Holisitic Molecule Interpretation"]:
    grade_list.append(grades)

conf_value_list = []
for values in df["CXSMILES's Confidence Levels"]:
    conf_value_list.append(values)

conf_classifications = []
for classifications in df["Confidence Classification"]:
    conf_classifications.append(classifications)

type_list = []
for types in df["Type of Mol"]
    type_list.append(types)


#Loads an empty Excel file named "canvas.xlsx" which will be transformed to a new Excel file for the specific category chosen below
workbook = load_workbook("DECIMER-Image-Segmentation/canvas.xlsx")
worksheet = workbook.active
workbook.save("DECIMER-Image-Segmentation/canvas.xlsx")

from openpyxl.drawing.image import Image

#Creates the Appropriate Column Titles for CXMolScribe Output
worksheet.cell(row=1,column=2).value = "File Path"
worksheet.cell(row=1,column=4).value = "Image From File Path"
worksheet.cell(row=1, column=5).value = "Image From Predicted SMILES"
worksheet.cell(row=1, column=6).value = "Predicted SMILES"
worksheet.cell(row=1,column=7).value = "Nuances in Image"
worksheet.cell(row=1,column=8).value = "Holistic Molecule Interpretation"
worksheet.cell(row=1,column=9).value = "CXSMILES's Confidence Levels"
worksheet.cell(row=1,column=10).value = "Confidence Classification"
worksheet.cell(row=1,column=11).value = "Type of Mol"

"""
Below is the function that will iterate through the types of molecular images present
and sort any relevant information regarding a specific classification into its own
Excel file.  In order to do this, change the moldesignation value to be equal to
any of the classification abbreviations seen in Table S1.  If the classification you
desire to extract has multiple abbreviations such as a Partial Segmentations having 
"ps" and "ps&es", have an expression such as this for the if condition:
if moldesignation == "ps" or moldesignation == "ps&es"
"""

row_value = 2
for digit,moldesignation in enumerate(df["Type of Mol"]):

    if moldesignation == "mol": #CHANGE THIS PER CLASSIFICATION
        
	#Adds appropriate information to the specific classification document
        
	worksheet.cell(row=row_value,column=2).value = file_paths[digit]
        dis_image = Image(file_paths[digit])
        dis_image.width = 62
        dis_image.height = 48
        worksheet.add_image(dis_image, "D"+str(row_value))
        #Adds values from above lists to their appropriate places in the specific classification document
        worksheet.cell(row=row_value,column=6).value = smiles_list[digit]
        worksheet.cell(row=row_value,column=7).value = nuance_list[digit]#Remove this if not dealing with "mol" classification
        worksheet.cell(row=row_value,column=8).value = grade_list[digit]
        worksheet.cell(row=row_value,column=9).value = conf_value_list[digit]
        worksheet.cell(row=row_value,column=10).value = conf_classifications[digit]
        worksheet.cell(row=row_value,column=11).value = moldesignation
        worksheet.row_dimensions[digit].height = 90
        worksheet.column_dimensions["D"].width = 70
        worksheet.column_dimensions["F"].width = 100
        worksheet.column_dimensions["G"].width = 40
        worksheet.column_dimensions["H"].width = 40
        worksheet.column_dimensions["I"].width = 50
        row_value+=1
        try:
            mol = Chem.MolFromSmiles(smiles_list[digit])
            smiles_path = str(smiles_list[digit])+".png"
            mol_image = Draw.MolsToImage([mol])
            mol_image.save(smiles_path)
            pred_smiles_image = Image(smiles_path)
            pred_smiles_image.width = 75
            pred_smiles_image.height = 70
            new_row = row_value - 1
            worksheet.add_image(pred_smiles_image, "E"+str(new_row))
        except:
            continue

#This will save the newly generated Excel file with the name you desire, change the name to fit your extracted classification
workbook.save("DECIMER-Image-Segmentation/CMAGE_mol.xlsx")#

print("Segmentation Classification Organization Completed")
