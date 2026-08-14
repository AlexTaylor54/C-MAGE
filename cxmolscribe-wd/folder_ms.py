import sys
import torch
import rdkit.Chem as Chem
from rdkit.Chem import Draw
import os
import pandas as pd
import numpy as np
import time
import cv2
import argparse
from pathlib import Path
from openpyxl import load_workbook
from huggingface_hub import hf_hub_download


#Every path defaults to a location derived from this file rather than from the working directory
CXMS_DIR = Path(__file__).resolve().parent
DIS_DIR = CXMS_DIR / "DECIMER-Image-Segmentation"

parser = argparse.ArgumentParser(
    description="Translate segmented structure images into SMILES with CXMolScribe.")
parser.add_argument("--results-excel", type=Path, default=DIS_DIR / "DIS_CMAGE_results.xlsx",
                    help="Spreadsheet of segment paths written by stage 2 (DECIMER).")
parser.add_argument("--output-dir", type=Path, default=DIS_DIR,
                    help="Directory for the classified images and completed spreadsheets.")
parser.add_argument("--canvas", type=Path, default=DIS_DIR / "canvas.xlsx",
                    help="Empty workbook used as the template for low-confidence results.")
parser.add_argument("--device", default=os.environ.get("CMAGE_DEVICE", "cpu"),
                    help="torch device for MolScribe: cpu (default), mps on Apple Silicon, "
                         "or cuda. Also read from CMAGE_DEVICE.")
args = parser.parse_args()

if not args.results_excel.is_file():
    raise SystemExit(
        f"Stage 2 results not found: {args.results_excel}\n"
        "Run pipeline_dis.py first, or pass --results-excel.")

os.makedirs(args.output_dir, exist_ok=True)

#Establishes model path for being run in this script
model_path = hf_hub_download('yujieq/MolScribe', 'swin_base_char_aux_1m.pth')

#Makes a dataframe out of the excel file created from the DECIMER-Image-Segmentation proccessing
df = pd.read_excel(args.results_excel)

#Takes DIS file paths from previous step and puts them in a list for CXMolScribe usage
file_paths = []
for paths in df["DIS Result File Paths"]:
    file_paths.append(paths)

#Loads the DECIMER-Image-Segmentation output excel as a workbook that can be edited
workbook = load_workbook(args.results_excel)
worksheet = workbook.active

#Loads an empty excel as a second workbook for low confidence values to be stored
dis_workbook = load_workbook(args.canvas)
dis_worksheet = dis_workbook.active

#Establishes the model to be used for translation
from openpyxl.drawing.image import Image
from molscribe import MolScribe
device = torch.device(args.device)
print("Device = " + str(device))
model = MolScribe(model_path, device)

#Creates the Appropriate Column Titles for CXMolScribe Output
worksheet.cell(row=1,column=2).value = "File Path"
worksheet.cell(row=1,column=4).value = "Image From File Path"
worksheet.cell(row=1, column=5).value = "Image From Predicted CXSMILES"
worksheet.cell(row=1, column=6).value = "Predicted CXSMILES"
worksheet.cell(row=1,column=7).value = "Nuances in Image"
worksheet.cell(row=1,column=8).value = "Holistic Molecule Interpretation"
worksheet.cell(row=1,column=9).value = "CXSMILES's Confidence Levels"
worksheet.cell(row=1,column=10).value = "Confidence Classification"
worksheet.cell(row=1,column=11).value = "Type of Mol"


dis_worksheet.cell(row=1,column=2).value = "File Path"
dis_worksheet.cell(row=1,column=4).value = "Image From File Path"
dis_worksheet.cell(row=1, column=5).value = "Image From Predicted CXSMILES"
dis_worksheet.cell(row=1, column=6).value = "Predicted CXSMILES"
dis_worksheet.cell(row=1,column=7).value = "Nuances in Image"
dis_worksheet.cell(row=1,column=8).value = "Holistic Molecule Interpretation"
dis_worksheet.cell(row=1,column=9).value = "CXSMILES's Confidence Levels"
dis_worksheet.cell(row=1,column=10).value = "Confidence Classification"
dis_worksheet.cell(row=1,column=11).value = "Type of Mol"


#Creates a list for predicted CXSMILES strings to be placed into
smiles_prediction = []


#Confidence Classifcation Strings and Counters
correct_classification = "High Confidence of SMILES being correct"
discard_classification = "Discard this SMILES prediction"

starttime= time.time()

correct_counter = 0
discard_counter = 0

#Establishes a folder path for images to be stored depending on their confidence classification
hc_folder = str(args.output_dir / "highconfidence_images")
dis_folder = str(args.output_dir / "lowconfidence_images")

#Creates sorted directories for rendered structure placement
os.makedirs(hc_folder, exist_ok=True)
os.makedirs(dis_folder, exist_ok=True)

hc_row_value = 2
dis_row_value = 2

#For loop to run each file path through CXMolScribe and organizing the results
for digit,fps in enumerate(file_paths):
    
    #Makes CXSMILES prediction of DECIMER-Image-Segmentation Image
    #Uses established Confidence Intervals to sort molecules based on CXMolScribe confidence
    #If prediction falls into the high confidence classification its information is added to the high confidence Excel and image folder
    #If prediction falls into the discard classification its information is added to the discard  Excel and image folder
    #Appropriate excel sheet variables are filled
    
    if digit == digit:
        prediction = model.predict_image_file(fps, return_atoms_bonds=True, return_confidence=True)
        smiles = prediction["smiles"]
        confidence = prediction["confidence"]
        
        #Retains the molecules figure connection within the naming of the CXSMILES translation
        cut_start = fps.index("_:") + 2
        cut_end = fps.index(".png")
        translation_identifier = fps[cut_start:cut_end]
        
        #CXMolScribe code for high confidnece translation
        if confidence >= 0.8431 and str(smiles) != "<invalid>":
            try:        
                worksheet.cell(row=hc_row_value, column=10).value = correct_classification
                correct_counter += 1

                #Appends molecule predictions to appropriate location on output spreadsheet
                worksheet.cell(row=hc_row_value, column=9).value = confidence
                worksheet.cell(row=hc_row_value, column=6).value = smiles
                smiles_prediction.append(smiles)

                #Creates an image of the ground truth DECIMER-Image-Segmentation output and organizes it into the excel sheet
                fp_image = Image(fps)
                fp_image.width = 62
                fp_image.height = 48
                worksheet.add_image(fp_image, "D"+str(hc_row_value))

                #Creates an image of the predicted CXSMILES string and organizes it into the excel sheet
                mol = Chem.MolFromSmiles(smiles)
                
                smiles_path = str(translation_identifier) + ".png"
                combined = os.path.join(hc_folder, smiles_path)
                final_smiles_path: str =  os.path.abspath(combined)
                
                mol_image = Draw.MolsToImage([mol])
                mol_image.save(final_smiles_path)
                
                pred_smiles_image = Image(final_smiles_path)
                pred_smiles_image.width = 75
                pred_smiles_image.height = 70
                worksheet.add_image(pred_smiles_image, "E"+str(hc_row_value))
                
                hc_row_value += 1
                
            except:
                hc_row_value += 1
                continue
        
        #CXMolScribe code for low confidence translations
        if confidence < 0.8431 and str(smiles) != "<invalid>":
            try:        
                
                #Adds the DECIMER-Image-Segmentation file path to the discard excel file
                dis_worksheet.cell(row=dis_row_value, column=2).value = fps

                dis_worksheet.cell(row=dis_row_value, column=10).value = discard_classification
                discard_counter += 1

                #Appends molecule predictions to appropriate location on output spreadsheet
                dis_worksheet.cell(row=dis_row_value, column=9).value = confidence
                dis_worksheet.cell(row=dis_row_value, column=6).value = smiles
                smiles_prediction.append(smiles)
                
                #Creates an image of the ground truth DECIMER-Image-Segmentation output and organizes it into the excel sheet
                fp_image = Image(fps)
                fp_image.width = 62
                fp_image.height = 48
                dis_worksheet.add_image(fp_image, "D"+str(dis_row_value))
                
                #Creates an image of the predicted CXSMILES string and organizes it into the excel sheet
                mol = Chem.MolFromSmiles(smiles)
                
                dis_smiles_path = str(translation_identifier) + ".png"
                dis_combined = os.path.join(dis_folder, dis_smiles_path)
                dis_final_smiles_path: str =  os.path.abspath(dis_combined)
                
                mol_image = Draw.MolsToImage([mol])
                mol_image.save(dis_final_smiles_path)
                
                pred_smiles_image = Image(dis_final_smiles_path)
                pred_smiles_image.width = 75
                pred_smiles_image.height = 70
                dis_worksheet.add_image(pred_smiles_image, "E"+str(dis_row_value))

                dis_row_value += 1
            except:
                dis_row_value += 1
                continue

        #CXMolScribe code for invalid translations
        if str(smiles) == "<invalid>":
            try:        
                
                #Adds the DECIMER-Image-Segmentation file path to the discard excel file
                dis_worksheet.cell(row=dis_row_value, column=2).value = fps

                dis_worksheet.cell(row=dis_row_value,column=10).value = discard_classification
                discard_counter += 1
                
                #Appends molecule predictions to appropriate location on output spreadsheet
                dis_worksheet.cell(row=dis_row_value, column=9).value = confidence
                dis_worksheet.cell(row=dis_row_value, column=6).value = smiles
                smiles_prediction.append(smiles)

                #Creates an image of the ground truth DECIMER-Image-Segmentation output and organizes it into the excel sheet
                fp_image = Image(fps)
                fp_image.width = 62
                fp_image.height = 48
                dis_worksheet.add_image(fp_image, "D"+str(dis_row_value))

                #Creates an image of the predicted SMILES string and organizes it into the excel sheet
                mol = Chem.MolFromSmiles(smiles)
                
                dis_smiles_path = str(translation_identifier) + ".png"
                dis_combined = os.path.join(dis_folder, dis_smiles_path)
                dis_final_smiles_path: str =  os.path.abspath(dis_combined)
                
                mol_image = Draw.MolsToImage([mol])
                mol_image.save(dis_final_smiles_path)
                
                pred_smiles_image = Image(dis_final_smiles_path)
                pred_smiles_image.width = 75
                pred_smiles_image.height = 70
                dis_worksheet.add_image(pred_smiles_image, "E"+str(dis_row_value))

                dis_row_value +=1
            
            except:
                dis_row_value += 1
                continue

        #Standardizes sizing of rows and columns in the excel sheets
        worksheet.row_dimensions[digit].height = 90
        worksheet.column_dimensions["D"].width = 70
        worksheet.column_dimensions["F"].width = 100
        worksheet.column_dimensions["G"].width = 40
        worksheet.column_dimensions["H"].width = 40
        worksheet.column_dimensions["I"].width = 50

        dis_worksheet.row_dimensions[digit].height = 90
        dis_worksheet.column_dimensions["D"].width = 70
        dis_worksheet.column_dimensions["F"].width = 100
        dis_worksheet.column_dimensions["G"].width = 40
        dis_worksheet.column_dimensions["H"].width = 40
        dis_worksheet.column_dimensions["I"].width = 50

endtime = time.time()
totaltime = endtime - starttime


print("Total CXMolScribe Time = " +str(totaltime)+ " seconds!")
print("High Confidence of Correct Counter = " +str(correct_counter))
print("Discard Counter = " +str(discard_counter))

#Saves Workbook as new excel file which now has all CXMolScribe results from above
workbook.save(args.output_dir / "Completed_HighConfidence_CMAGE.xlsx")
dis_workbook.save(args.output_dir / "Completed_LowConfidence_CMAGE.xlsx")

print("Whole Pipeline Extracted!")
